"""Local Excel tools (openpyxl). No authentication required."""

from __future__ import annotations

import os
from typing import Any

import openpyxl


def read_excel(path: str, sheet: str | None = None) -> dict[str, Any]:
    """Read a local .xlsx file and return its rows.

    The first row is treated as a header when every cell in it is a
    non-empty string; otherwise rows are returned positionally.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"No such spreadsheet: {path}")

    book = openpyxl.load_workbook(path, data_only=True)
    ws = book[sheet] if sheet else book.active

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # Drop trailing rows that openpyxl pads onto sparse sheets.
    while rows and all(c is None for c in rows[-1]):
        rows.pop()

    if not rows:
        return {"path": path, "sheet": ws.title, "headers": [], "rows": []}

    first = rows[0]
    has_header = bool(first) and all(isinstance(c, str) and c.strip() for c in first)
    headers = [str(c) for c in first] if has_header else []
    body = rows[1:] if has_header else rows

    return {
        "path": path,
        "sheet": ws.title,
        "headers": headers,
        "row_count": len(body),
        "rows": [dict(zip(headers, r)) if has_header else list(r) for r in body],
    }


def write_excel(
    path: str, rows: list[list[Any]], sheet: str | None = None
) -> dict[str, Any]:
    """Write `rows` to a local .xlsx file, creating or replacing the sheet."""
    if os.path.exists(path):
        book = openpyxl.load_workbook(path)
        if sheet and sheet in book.sheetnames:
            del book[sheet]
        ws = book.create_sheet(sheet) if sheet else book.active
        if not sheet:
            ws.delete_rows(1, ws.max_row)
    else:
        book = openpyxl.Workbook()
        ws = book.active
        if sheet:
            ws.title = sheet

    for row in rows:
        ws.append(list(row))
    book.save(path)
    return {"path": path, "sheet": ws.title, "rows_written": len(rows)}


EXCEL_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_excel",
            "description": (
                "Read a local Excel (.xlsx) file from disk and return its rows. "
                "Use this for any local spreadsheet file; it needs no credentials."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to the .xlsx file, e.g. 'test.xlsx'.",
                    },
                    "sheet": {
                        "type": "string",
                        "description": "Optional worksheet name. Defaults to the active sheet.",
                    },
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_excel",
            "description": (
                "Write rows to a local Excel (.xlsx) file, replacing existing "
                "contents of the target sheet. This overwrites data."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Destination .xlsx path."},
                    "rows": {
                        "type": "array",
                        "description": "Rows to write; each row is an array of cell values.",
                        "items": {"type": "array", "items": {}},
                    },
                    "sheet": {
                        "type": "string",
                        "description": "Optional worksheet name.",
                    },
                },
                "required": ["path", "rows"],
            },
        },
    },
]

EXCEL_FUNCS = {"read_excel": read_excel, "write_excel": write_excel}
