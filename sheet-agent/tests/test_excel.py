"""Local Excel tool tests against real files on disk."""

from __future__ import annotations

import openpyxl
import pytest

from tools.excel import read_excel, write_excel


def test_read_excel_parses_headers_and_rows(baseline_xlsx):
    data = read_excel(str(baseline_xlsx))
    assert data["headers"] == ["Name", "Amount"]
    assert data["row_count"] == 3
    assert data["rows"][1] == {"Name": "Bob", "Amount": 340}


def test_read_excel_missing_file_raises():
    with pytest.raises(FileNotFoundError):
        read_excel("definitely-not-here.xlsx")


def test_read_excel_named_sheet(tmp_path):
    book = openpyxl.Workbook()
    book.active.title = "First"
    book.active.append(["a"])
    second = book.create_sheet("Second")
    second.append(["Col"])
    second.append(["value"])
    path = tmp_path / "multi.xlsx"
    book.save(path)

    data = read_excel(str(path), sheet="Second")
    assert data["sheet"] == "Second"
    assert data["rows"] == [{"Col": "value"}]


def test_read_excel_without_header_row(tmp_path):
    """A numeric first row is data, not a header."""
    book = openpyxl.Workbook()
    book.active.append([1, 2])
    book.active.append([3, 4])
    path = tmp_path / "nohdr.xlsx"
    book.save(path)

    data = read_excel(str(path))
    assert data["headers"] == []
    assert data["rows"] == [[1, 2], [3, 4]]


def test_read_excel_empty_sheet(tmp_path):
    path = tmp_path / "empty.xlsx"
    openpyxl.Workbook().save(path)
    data = read_excel(str(path))
    assert data["rows"] == []


def test_write_then_read_roundtrip(tmp_path):
    path = tmp_path / "out.xlsx"
    result = write_excel(str(path), [["Name", "Qty"], ["Widget", 7]])
    assert result["rows_written"] == 2

    data = read_excel(str(path))
    assert data["rows"] == [{"Name": "Widget", "Qty": 7}]


def test_write_excel_replaces_existing_contents(baseline_xlsx):
    """Writing overwrites rather than appending -- the documented behaviour."""
    write_excel(str(baseline_xlsx), [["Name", "Amount"], ["Zoe", 5]])
    data = read_excel(str(baseline_xlsx))
    assert data["row_count"] == 1
    assert data["rows"] == [{"Name": "Zoe", "Amount": 5}]
